# clients/views.py
from django.shortcuts import get_object_or_404
from django.db.models import Q
from django.utils import timezone
from rest_framework import viewsets, status, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.exceptions import ValidationError
from django_filters.rest_framework import DjangoFilterBackend
import logging

from .models import Client
from .serializers import (
    ClientSerializer, 
    ClientCreateSerializer,
    ClientUpdateSerializer,
    ClientDetailSerializer,
    ClientExportSerializer
)
from .permissions import IsAdminOrManagerOrReadOnly, IsAdminOrManagerOrOwnerReadOnly
from .filters import ClientFilter
from .pagination import StandardResultsSetPagination

logger = logging.getLogger(__name__)

class ClientViewSet(viewsets.ModelViewSet):
    """
    API endpoint for managing clients with advanced features.
    """
    queryset = Client.objects.all().select_related('user_account')
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrManagerOrReadOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ClientFilter
    search_fields = ['name', 'contact_email', 'company_name', 'contact_person', 'phone']
    ordering_fields = ['name', 'company_name', 'created_at', 'updated_at']
    ordering = ['name']
    
    # Different parsers for different content types
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get_serializer_class(self):
        """
        Return appropriate serializer class based on action.
        """
        if self.action == 'create':
            return ClientCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return ClientUpdateSerializer
        elif self.action == 'retrieve':
            return ClientDetailSerializer
        elif self.action == 'export':
            return ClientExportSerializer
        return ClientSerializer
    
    def get_queryset(self):
        """
        Filter queryset based on user permissions and query parameters.
        """
        queryset = super().get_queryset()
        
        # If user is not admin or manager, apply restrictions
        if self.request.user.role not in ['admin', 'manager']:
            # If user has a client profile, show only that
            if hasattr(self.request.user, 'client_profile'):
                queryset = queryset.filter(id=self.request.user.client_profile.id)
            else:
                # Non-admin/manager users without client profile see no clients
                queryset = queryset.none()
        
        # Filter by company if company_id provided
        company_id = self.request.query_params.get('company_id', None)
        if company_id:
            queryset = queryset.filter(company_id=company_id)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        # Filter by creation date range
        created_after = self.request.query_params.get('created_after', None)
        created_before = self.request.query_params.get('created_before', None)
        if created_after:
            queryset = queryset.filter(created_at__gte=created_after)
        if created_before:
            queryset = queryset.filter(created_at__lte=created_before)
        
        return queryset
    
    def perform_create(self, serializer):
        """
        Custom create logic with audit trail.
        """
        try:
            # Add created_by user to the instance
            instance = serializer.save()
            
            # Log the creation
            logger.info(
                f"Client created: {instance.name} (ID: {instance.id}) "
                f"by user: {self.request.user.email}"
            )
            
            # Send notification if needed
            self._send_client_creation_notification(instance)
            
        except Exception as e:
            logger.error(f"Error creating client: {str(e)}")
            raise ValidationError(f"Failed to create client: {str(e)}")
    
    def perform_update(self, serializer):
        """
        Custom update logic with audit trail.
        """
        try:
            instance = serializer.save()
            
            # Log the update
            logger.info(
                f"Client updated: {instance.name} (ID: {instance.id}) "
                f"by user: {self.request.user.email}"
            )
            
            # Create audit trail entry
            self._create_audit_trail(instance, 'UPDATE')
            
        except Exception as e:
            logger.error(f"Error updating client: {str(e)}")
            raise ValidationError(f"Failed to update client: {str(e)}")
    
    def perform_destroy(self, instance):
        """
        Soft delete implementation instead of hard delete.
        """
        try:
            # Instead of deleting, mark as inactive and archive
            instance.is_active = False
            instance.archived_at = timezone.now()
            instance.archived_by = self.request.user
            instance.save()
            
            # Log the deletion
            logger.info(
                f"Client archived: {instance.name} (ID: {instance.id}) "
                f"by user: {self.request.user.email}"
            )
            
            # Create audit trail entry
            self._create_audit_trail(instance, 'ARCHIVE')
            
        except Exception as e:
            logger.error(f"Error archiving client: {str(e)}")
            raise ValidationError(f"Failed to archive client: {str(e)}")
    
    @action(detail=False, methods=['get'], url_path='search')
    def search_clients(self, request):
        """
        Advanced search endpoint with multiple criteria.
        """
        query = request.query_params.get('q', '')
        
        if not query:
            return Response(
                {'detail': 'Search query parameter "q" is required.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Complex search across multiple fields
            search_results = Client.objects.filter(
                Q(name__icontains=query) |
                Q(contact_email__icontains=query) |
                Q(company_name__icontains=query) |
                Q(contact_person__icontains=query) |
                Q(phone__icontains=query) |
                Q(address__icontains=query) |
                Q(notes__icontains=query)
            ).distinct()
            
            # Apply permissions - only restrict for non-admin/manager users
            if request.user.role not in ['admin', 'manager']:
                search_results = search_results.filter(user_account=request.user)
            
            # Paginate results
            page = self.paginate_queryset(search_results)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
            
            serializer = self.get_serializer(search_results, many=True)
            return Response(serializer.data)
            
        except Exception as e:
            logger.error(f"Error searching clients: {str(e)}")
            return Response(
                {'detail': 'An error occurred during search.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='stats')
    def client_statistics(self, request):
        """
        Get client statistics for dashboard.
        """
        try:
            total_clients = Client.objects.count()
            active_clients = Client.objects.filter(is_active=True).count()
            clients_with_users = Client.objects.filter(user_account__isnull=False).count()
            recent_clients = Client.objects.filter(
                created_at__gte=timezone.now() - timezone.timedelta(days=30)
            ).count()
            
            # Company-wise breakdown
            company_stats = Client.objects.values('company_name').annotate(
                count=models.Count('id'),
                latest=models.Max('created_at')
            ).order_by('-count')[:10]
            
            return Response({
                'total_clients': total_clients,
                'active_clients': active_clients,
                'clients_with_user_accounts': clients_with_users,
                'clients_last_30_days': recent_clients,
                'top_companies': list(company_stats)
            })
            
        except Exception as e:
            logger.error(f"Error generating client statistics: {str(e)}")
            return Response(
                {'detail': 'Failed to generate statistics.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create_clients(self, request):
        """
        Create multiple clients at once (CSV/JSON import).
        """
        if request.user.role not in ['admin', 'manager']:
            return Response(
                {'detail': 'Permission denied.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        data = request.data
        if not isinstance(data, list):
            return Response(
                {'detail': 'Expected a list of clients.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_clients = []
        errors = []
        
        for idx, client_data in enumerate(data):
            try:
                serializer = ClientCreateSerializer(data=client_data)
                if serializer.is_valid():
                    client = serializer.save()
                    created_clients.append(client.id)
                else:
                    errors.append({
                        'index': idx,
                        'data': client_data,
                        'errors': serializer.errors
                    })
            except Exception as e:
                errors.append({
                    'index': idx,
                    'data': client_data,
                    'errors': str(e)
                })
        
        return Response({
            'created': len(created_clients),
            'failed': len(errors),
            'created_ids': created_clients,
            'errors': errors
        })
    
    @action(detail=True, methods=['post'], url_path='upload-document')
    def upload_document(self, request, pk=None):
        """
        Upload document for a specific client.
        """
        client = self.get_object()
        
        if 'document' not in request.FILES:
            return Response(
                {'detail': 'No document file provided.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        document_file = request.FILES['document']
        
        # Validate file size (max 5MB)
        if document_file.size > 5 * 1024 * 1024:
            return Response(
                {'detail': 'File size exceeds 5MB limit.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file type
        allowed_types = ['application/pdf', 'image/jpeg', 'image/png', 'application/msword']
        if document_file.content_type not in allowed_types:
            return Response(
                {'detail': 'Invalid file type. Allowed: PDF, JPEG, PNG, DOC.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Save document to client's storage
            # You would integrate with your storage solution here
            # Example: document = ClientDocument.objects.create(...)
            
            logger.info(
                f"Document uploaded for client {client.name}: "
                f"{document_file.name} by {request.user.email}"
            )
            
            return Response({
                'detail': 'Document uploaded successfully.',
                'filename': document_file.name,
                'size': document_file.size,
                'type': document_file.content_type
            })
            
        except Exception as e:
            logger.error(f"Error uploading document: {str(e)}")
            return Response(
                {'detail': 'Failed to upload document.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='export')
    def export_clients(self, request):
        """
        Export clients data in various formats.
        """
        format_type = request.query_params.get('format', 'json').lower()
        
        if format_type not in ['json', 'csv', 'xlsx']:
            return Response(
                {'detail': 'Invalid format. Supported: json, csv, xlsx'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        queryset = self.filter_queryset(self.get_queryset())
        
        if format_type == 'json':
            serializer = self.get_serializer(queryset, many=True)
            return Response(serializer.data)
        
        elif format_type == 'csv':
            # Generate CSV response
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="clients.csv"'
            
            writer = csv.writer(response)
            # Write header
            writer.writerow([
                'ID', 'Name', 'Email', 'Phone', 'Company', 
                'Contact Person', 'Created At', 'Updated At'
            ])
            
            # Write data
            for client in queryset:
                writer.writerow([
                    client.id,
                    client.name,
                    client.contact_email or '',
                    client.phone or '',
                    client.company_name or '',
                    client.contact_person or '',
                    client.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    client.updated_at.strftime('%Y-%m-%d %H:%M:%S')
                ])
            
            return response
        
        else:  # xlsx
            # Generate Excel file
            import openpyxl
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clients"
            
            # Write header
            headers = ['ID', 'Name', 'Email', 'Phone', 'Company', 
                      'Contact Person', 'Address', 'Notes', 'Created At', 'Updated At']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, client in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=client.id)
                ws.cell(row=row, column=2, value=client.name)
                ws.cell(row=row, column=3, value=client.contact_email or '')
                ws.cell(row=row, column=4, value=client.phone or '')
                ws.cell(row=row, column=5, value=client.company_name or '')
                ws.cell(row=row, column=6, value=client.contact_person or '')
                ws.cell(row=row, column=7, value=client.address or '')
                ws.cell(row=row, column=8, value=client.notes or '')
                ws.cell(row=row, column=9, value=client.created_at)
                ws.cell(row=row, column=10, value=client.updated_at)
            
            # Create HTTP response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="clients.xlsx"'
            wb.save(response)
            
            return response
    
    @action(detail=True, methods=['get'], url_path='activities')
    def client_activities(self, request, pk=None):
        """
        Get recent activities for a specific client.
        """
        client = self.get_object()
        
        # This would typically come from an AuditLog model
        # For now, return mock data
        activities = [
            {
                'id': 1,
                'action': 'CREATED',
                'timestamp': client.created_at.isoformat(),
                'user': 'System' if not client.user_account else client.user_account.email,
                'details': 'Client record created'
            },
            {
                'id': 2,
                'action': 'UPDATED',
                'timestamp': client.updated_at.isoformat(),
                'user': request.user.email,
                'details': 'Client information updated'
            }
        ]
        
        return Response({
            'client_id': client.id,
            'client_name': client.name,
            'activities': activities
        })
    
    def _send_client_creation_notification(self, client):
        """
        Send notification about new client creation.
        In production, integrate with your notification system.
        """
        # Example: Send email to admin
        # from django.core.mail import send_mail
        # send_mail(
        #     f'New Client Created: {client.name}',
        #     f'A new client has been created:\n\nName: {client.name}\nEmail: {client.contact_email}',
        #     'noreply@yourdomain.com',
        #     ['admin@yourdomain.com'],
        #     fail_silently=True,
        # )
        pass
    
    def _create_audit_trail(self, instance, action):
        """
        Create audit trail entry for client actions.
        """
        # In production, you would save to an AuditLog model
        # Example:
        # AuditLog.objects.create(
        #     user=self.request.user,
        #     action=action,
        #     model_name='Client',
        #     object_id=instance.id,
        #     object_repr=str(instance),
        #     changes=self._get_changes(instance)  # You'd need to track changes
        # )
        pass